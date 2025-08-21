import cv2
import numpy as np
import os
import csv
import pandas as pd
from openpyxl import load_workbook


"""
classificação de tamanho de poros:
    megapore
        large, medium, small
    mesopore
        very large, large, medium, small very small
    micropore
        
"""

def calculate_media(image, file_name, output_folder_path):
    # Conversão de medidas da imagem
    image_width_mm = 26
    image_height_mm = 18
    image_width_px = 15191
    image_height_px = 10692
    pixel_area_mm2 = (image_width_mm * image_height_mm) / (image_width_px * image_height_px)

    area_minima_mm2 = 0.005
    LIMIAR_PEQUENO = 0.0625
    LIMIAR_MEDIO = 4

    """
    mvs = mesopore very small -> 0,0625 - 0,25mm
    ms = mesopore small -> 0,25 - 0,5mm
    mm = mesopore medium -> 0,25 - 1mm
    ml = mesopore large -> 1 - 2mm
    mvl = mesopore very large -> 2 - 4mm
    
    megaS = megapore small -> 4 - 16mm
    megaM = megapore medium -> 16 - 32mm
    megaL = megapore large -> 32 - 256mm
    
    micropore < 0,0625
    """

    tamanhos = {
        'micropore': [],
        'mesopore': [],
        'megapore': [],
    }

    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    lower_blue = np.array([90, 50, 50])
    upper_blue = np.array([130, 255, 255])
    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    bw_image = cv2.bitwise_and(image, image, mask=mask)
    bw_gray = cv2.cvtColor(bw_image, cv2.COLOR_BGR2GRAY)
    _, bw_binary = cv2.threshold(bw_gray, 1, 255, cv2.THRESH_BINARY)

    contours, _ = cv2.findContours(bw_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    output_base = np.zeros((bw_binary.shape[0], bw_binary.shape[1], 3), dtype=np.uint8)
    output_base[bw_binary == 255] = (255, 255, 255)

    mask_circles = np.zeros_like(bw_binary)
    circle_areas_mm2 = []


    for cnt in contours:
        area_px = cv2.contourArea(cnt)
        area_mm2 = area_px * pixel_area_mm2

        if area_mm2 >= area_minima_mm2:
            (x, y), radius = cv2.minEnclosingCircle(cnt)
            center = (int(x), int(y))
            radius = int(radius)

            if area_mm2 < LIMIAR_PEQUENO:
                color = (255, 0, 0)
                tamanhos['micropore'].append(area_mm2)
            elif area_mm2 < LIMIAR_MEDIO:
                color = (0, 255, 255)
                tamanhos['mesopore'].append(area_mm2)
            else:
                color = (0, 0, 255)
                tamanhos['megapore'].append(area_mm2)

            cv2.circle(output_base, center, radius, color, 15)
            cv2.circle(mask_circles, center, radius, 255, -1)

            circle_area_px = np.pi * (radius ** 2)
            circle_areas_mm2.append(circle_area_px * pixel_area_mm2)

    pores_within_circles = cv2.bitwise_and(bw_binary, mask_circles)
    white_pixels = cv2.countNonZero(pores_within_circles)
    total_real_pore_area_mm2 = white_pixels * pixel_area_mm2

    valid_pores = len(circle_areas_mm2)
    mean_real_area = total_real_pore_area_mm2 / valid_pores if valid_pores > 0 else 0

    #quantidade de poros encontrados = valid_pores
    #area medial real em pixels = mean_real_area
    #area minima = area_minima_mm2

    # --- SALVAR DADOS EM EXCEL ---
    try:
        workbook = load_workbook(filename="planilhaModelo.xlsx")
        print("Arquivo aberto com sucesso!")

        # Acessar a primeira planilha
        sheet = workbook['dados']
        linha_C = linha_D = linha_E = linha_F = linha_G = linha_H = linha_I = linha_J = 4  # começa na linha 4 para Mesopore Very Small

        # coloca os valores na planilha

        for i, valor in enumerate(tamanhos['micropore']):
            sheet.cell(row=i + 4, column=2, value=valor)

        for valor in tamanhos['mesopore']:
            if valor < 0.25:
                sheet.cell(row=linha_C, column=3, value=valor)
                linha_C += 1
            elif valor < 0.5:
                sheet.cell(row=linha_D, column=4, value=valor)
                linha_D += 1
            elif valor < 1:
                sheet.cell(row=linha_E, column=5, value=valor)
                linha_E += 1
            elif valor < 2:
                sheet.cell(row=linha_F, column=6, value=valor)
                linha_F += 1
            elif valor < 4:
                sheet.cell(row=linha_G, column=7, value=valor)
                linha_G += 1

        for valor in tamanhos['megapore']:
            if valor < 16:
                sheet.cell(row=linha_H, column=8, value=valor)
                linha_H += 1
            elif valor < 32:
                sheet.cell(row=linha_I, column=9, value=valor)
                linha_I += 1
            elif valor < 256:
                sheet.cell(row=linha_J, column=10, value=valor)
                linha_J += 1

        # salva planilha nova
        import os
        # Substitua pelo caminho real
        pastaExcel = "tabelasExcel"
        caminho_completo = os.path.join(output_folder_path, pastaExcel)

        try:
            os.makedirs(caminho_completo)
            print(f"Pasta '{pastaExcel}' criada com sucesso em '{output_folder_path}'")
        except FileExistsError:
            print(f"A pasta '{pastaExcel}' já existe em '{output_folder_path}'")
        except Exception as e:
            print(f"Ocorreu um erro ao criar a pasta: {e}")

        workbook.save(output_folder_path+'/'+pastaExcel+'/'+file_name+"planilhaModelo1.xlsx")

    except FileNotFoundError:
        print("Erro: Arquivo não encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

    # --- SALVAR IMAGEM ---
    output_image_path = "poros_classificados.jpg"
    cv2.imwrite(output_image_path, output_base)
    print(file_name)
    print(f"CSV salvo como: ")
    print(f"Imagem com poros classificados salva como: {output_image_path}")

    return mean_real_area, output_base


def calculate_porosity(image):
    ycbcr = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
    cb_channel = ycbcr[:, :, 2]
    _, BW_otsu = cv2.threshold(cb_channel, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    inverted_BW = cv2.bitwise_not(BW_otsu)
    total_area = BW_otsu.size
    white_area = np.sum(inverted_BW == 255)
    porosity = (white_area / total_area) * 100
    return porosity, inverted_BW


def crop_and_calculate_porosity(input_folder_path, output_folder_path, top_border, bottom_border, left_border, right_border, progress_bar):
    if not os.path.exists(output_folder_path):
        os.makedirs(output_folder_path)

    file_list = [f for f in os.listdir(input_folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff'))]
    total_files = len(file_list)
    results = []

    for idx, file_name in enumerate(file_list):
        image_path = os.path.join(input_folder_path, file_name)
        img = cv2.imread(image_path)

        if img is None:
            print(f"Não foi possível carregar a imagem {file_name}")
            continue

        height, width, _ = img.shape
        start_y = top_border
        end_y = height - bottom_border
        start_x = left_border
        end_x = width - right_border

        if start_y >= end_y or start_x >= end_x:
            print(f"As dimensões de corte excedem o tamanho da imagem {file_name}")
            continue

        # Cortar imagem
        cropped_image = img[start_y:end_y, start_x:end_x]
        new_file_name = f"{os.path.splitext(file_name)[0]}_cropped{os.path.splitext(file_name)[1]}"
        new_image_path = os.path.join(output_folder_path, new_file_name)
        cv2.imwrite(new_image_path, cropped_image)

        # Porosidade (por porcentagem)
        porosity, inverted_BW = calculate_porosity(cropped_image)
        output_otsu_name = os.path.splitext(new_file_name)[0] + '_otsu.png'
        output_otsu_path = os.path.join(output_folder_path, output_otsu_name)
        cv2.imwrite(output_otsu_path, inverted_BW)

        # Cálculo de áreas e classificação dos poros
        mean_pore_area, classified_img = calculate_media(cropped_image, file_name, output_folder_path)
        output_poros_name = os.path.splitext(new_file_name)[0] + '_pores_classified.png'
        output_poros_path = os.path.join(output_folder_path, output_poros_name)
        cv2.imwrite(output_poros_path, classified_img)

        results.append({
            'Imagem': new_file_name,
            'Porosidade (%)': porosity,
            'Média dos poros (mm²)': mean_pore_area
        })

        progress_bar["value"] = (idx + 1) / total_files * 100
        progress_bar.update_idletasks()

    df = pd.DataFrame(results)
    return df

