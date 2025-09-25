import csv
import openpyxl
from numpy.ma.extras import column_stack
from openpyxl import load_workbook
import random
import pandas as pd
from analiseEstatistica import valor_medio, desvio_padrao, erro_padrao, incerteza, predominante
import win32com.client as win32

def testeMedia(tamanhos,nome,caminho):

    # Substitua 'nome_do_arquivo.xlsx' pelo caminho do seu arquivo
    try:
        workbook = load_workbook(filename="planilhaModelo.xlsx")
        print("Arquivo aberto com sucesso!")

        # Acessar a primeira planilha
        sheet = workbook['dados']
        linha_C = linha_D = linha_E = linha_F = linha_G = linha_H = linha_I = linha_J = 4  # começa na linha 4 para Mesopore Very Small

        #coloca os valores na planilha

        for i, valor in enumerate(tamanhos['micropore']):
            if valor < 0.0625:
                sheet.cell(row=i + 4, column=2, value=valor)

        for valor in tamanhos['mesopore']:
            if valor < 0.0625:
                sheet.cell(row=linha_C, column=3, value=valor)
                linha_C += 1
            elif valor < 0.5:
                sheet.cell(row=linha_D, column=4, value=valor)
                linha_D += 1
            elif valor < 1:
                sheet.cell(row=linha_E, column=5, value=valor)
                linha_E += 1
            elif valor < 2:
                sheet.cell(row=linha_F, column=6, value=valor)
                linha_F += 1
            elif valor < 4:
                sheet.cell(row=linha_G, column=7, value=valor)
                linha_G += 1

        for valor in tamanhos['megapore']:
            if valor < 12:
                sheet.cell(row=linha_H, column=8, value=valor)
                linha_H += 1
            elif valor < 13:
                sheet.cell(row=linha_I, column=9, value=valor)
                linha_I += 1
            elif valor < 14:
                sheet.cell(row=linha_J, column=10, value=valor)
                linha_J += 1


        # salva planilha nova
        import os
        # Substitua pelo caminho real
        pastaExcel = "tabelasExcel"
        caminho_completo = os.path.join(caminho, pastaExcel)

        try:
            os.makedirs(caminho_completo)
            print(f"Pasta '{pastaExcel}' criada com sucesso em '{caminho}'")
        except FileExistsError:
            print(f"A pasta '{pastaExcel}' já existe em '{caminho}'")
        except Exception as e:
            print(f"Ocorreu um erro ao criar a pasta: {e}")

        caminho_dados = caminho + '/' + pastaExcel + '/' + nome + ".xlsx"
        workbook.save(caminho_dados)
        if caminho_dados:
            testeEst(caminho_dados)

    except FileNotFoundError:
        print("Erro: Arquivo não encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


def testeEst(caminho):
    # Carregar a planilha
    print('entrou', caminho)
    workbook = load_workbook(caminho)
    sheet = workbook['dados'] # ou workbook["NomeDaAba"]

    # Pegar todas as linhas a partir da linha 4
    micropore = {
        'micropore': []
    }
    mesopore = {
        'mesoporeVerySmall': [],
        'mesoporeSmall': [],
        'mesoporeMedium': [],
        'mesoporeLarge': [],
        'mesoporeVeryLarge': []
    }
    megapore = {
        'megaporeSmall': [],
        'megaporeMedium': [],
        'megaporeLarge': []
    }


    #dados do micropore
    for row in sheet.iter_rows(min_row=4, min_col=2, max_col=2, values_only=True):
        micropore['micropore'].append(list(row))
    micropore = [row[0] for row in micropore['micropore'] if row[0] is not None]


    #DADOS MESOPORE
    for row in sheet.iter_rows(min_row=4, min_col=3, max_col=3, values_only=True):
        mesopore['mesoporeVerySmall'].append(list(row))
    mesoporeVerySmall = [row[0] for row in mesopore['mesoporeVerySmall'] if row[0] is not None]


    for row in sheet.iter_rows(min_row=4, min_col=4, max_col=4, values_only=True):
        mesopore['mesoporeSmall'].append(list(row))
    mesoporeSmall = [row[0] for row in mesopore['mesoporeSmall'] if row[0] is not None]

    for row in sheet.iter_rows(min_row=4, min_col=5, max_col=5, values_only=True):
        mesopore['mesoporeMedium'].append(list(row))
    mesoporeMedium = [row[0] for row in mesopore['mesoporeMedium'] if row[0] is not None]

    for row in sheet.iter_rows(min_row=4, min_col=6, max_col=6, values_only=True):
        mesopore['mesoporeLarge'].append(list(row))
    mesoporeLarge = [row[0] for row in mesopore['mesoporeLarge'] if row[0] is not None]

    for row in sheet.iter_rows(min_row=4, min_col=7, max_col=7, values_only=True):
        mesopore['mesoporeVeryLarge'].append(list(row))
    mesoporeVeryLarge = [row[0] for row in mesopore['mesoporeVeryLarge'] if row[0] is not None]

    mesopore_all = [valor[0] for lista in mesopore.values() for valor in lista if valor[0] is not None]


    #DADOS MEGAPORE
    for row in sheet.iter_rows(min_row=4, min_col=8, max_col=8, values_only=True):
        megapore['megaporeSmall'].append(list(row))
    megaporeSmall = [row[0] for row in megapore['megaporeSmall'] if row[0] is not None]

    for row in sheet.iter_rows(min_row=4, min_col=9, max_col=9, values_only=True):
        megapore['megaporeMedium'].append(list(row))
    megaporeMedium = [row[0] for row in megapore['megaporeMedium'] if row[0] is not None]

    for row in sheet.iter_rows(min_row=4, min_col=10, max_col=10, values_only=True):
        megapore['megaporeLarge'].append(list(row))
    megaporeLarge = [row[0] for row in megapore['megaporeLarge'] if row[0] is not None]

    megapore_all = [valor[0] for lista in megapore.values() for valor in lista if valor[0] is not None]

    pores = micropore + megapore_all + megapore_all

    valores = micropore, mesoporeVerySmall, mesoporeSmall, mesoporeMedium, mesoporeLarge, mesoporeVeryLarge, mesopore_all, megaporeSmall, megaporeMedium, megaporeLarge, megapore_all, pores
    testeatualizaExcel(valores, caminho)

def testeatualizaExcel(valores, caminho):
    print('entrou2')
    wb = load_workbook(caminho)
    aba = wb['analise']  # ou wb["NomeDaAba"]


    #atualiza com valor médio
    for j, valor in enumerate(valores, start=1):
        if valor:  # se não for lista vazia
            media = valor_medio(valor)
        else:
            media = "N/A"   # pode ser 0, None ou "N/A", depende do que você quer na planilha
        aba.cell(row=4, column=j+1, value=media)

    for j, valor in enumerate(valores, start=1):
        if valor:  # se não for lista vazia
            dp = desvio_padrao(valor)
        else:
            dp = "N/A"   # pode ser 0, None ou "N/A", depende do que você quer na planilha
        aba.cell(row=5, column=j+1, value=dp)

    for j, valor in enumerate(valores, start=1):
        if valor:  # se não for lista vazia
            ep = erro_padrao(valor)
        else:
            ep = "N/A"   # pode ser 0, None ou "N/A", depende do que você quer na planilha
        aba.cell(row=6, column=j+1, value=ep)

    for j, valor in enumerate(valores, start=1):
        if valor:  # se não for lista vazia
            inc = incerteza(valor)
        else:
            inc = "N/A"   # pode ser 0, None ou "N/A", depende do que você quer na planilha
        aba.cell(row=7, column=j+1, value=inc)

    pre = predominante(valores)
     # pode ser 0, None ou "N/A", depende do que você quer na planilha
    aba.cell(row=8, column=2, value=pre)

    wb.save(caminho)  # ou outro nome para não sobrescrever

def teste_histograma():
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True  # mostra o Excel

    wb = excel.Workbooks.Open(r'C:\Users\JuliaBarbosa\Downloads\Processando-LaminasN\planilhaModelo1.xlsx')
    wn = wb.Sheets('analise')
    ws = wb.Sheets('dados')

    dados_range = ws.Range('B4:B11')  # ajuste conforme seus dados

    # Criar histograma
    #AddChart2(Style, Type, Left, Top, Width, Height)
    chart = wn.Shapes.AddChart2(201, 51, 300, 140, 500, 300)  # 201=histograma, 51=coluna
    chart.Chart.SetSourceData(dados_range)

    # Salvar e fechar
    wb.Save()
    wb.Close()
    excel.Quit()

def teste(valores):
    print(predominante(valores))






























