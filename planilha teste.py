from openpyxl import Workbook

# Cria uma nova planilha
wb = Workbook()

# Seleciona a aba ativa
ws = wb.active

# Adiciona um valor na célula A1
ws['A1'] = 'Olá, Mundo!'

# Adiciona um valor na célula B2
ws.cell(row=4, column=2, value=123)

# Salva a planilha
wb.save("planilha_exemplo.xlsx")